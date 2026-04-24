import pandas as pd
import os
from datetime import datetime
from shutil import copyfile
import openpyxl
import re
import sys
import unicodedata

def resource_path(relative_path):
    """
    Lấy đường dẫn file:
    - Khi chạy .py → lấy thư mục code
    - Khi chạy .exe → lấy thư mục _MEIPASS
    """
    try:
        base_path = sys._MEIPASS  # PyInstaller
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# ===== KIỂM TRA FILE / SHEET / CỘT ===== CHÍNH ====
# FILE_CONFIG = {
#     "file1": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\22. BAO CAO MMK\DU LIEU\T3\thoi_gian_gia_cong.xlsx",
#         "sheets": {
#             "Thời gian gia công": {"min_cols": 12}
#         }
#     },
#     "file2": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\22. BAO CAO MMK\DU LIEU\T3\Download1_2.xlsx",
#         "sheets": {
#             "Tỉ lệ hoạt động máy": {"min_cols": 3},
#             "BV,PCS HT trong tháng theo CĐ": {"min_cols": 3},
#             "Nhận BV,PCS trong tháng": {"min_cols": 5},
#         }
#     },
#     "file3": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\22. BAO CAO MMK\DU LIEU\T3\Download2_2026-02-26_2026-03-30.xlsx",
#         "sheets": {
#             "Số tiền hoàn thành": {"min_cols": 44},
#             "Tuân Thủ Kì Hạn": {"min_cols": 18},
#         }
#     },
#     "file4": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\22. BAO CAO MMK\DU LIEU\T3\Download3_2026-02-26 07_00_00_2026-03-30 06_59_00.xlsx",
#         "sheets": {
#             "BV.PCS Tồn đọng": {"min_cols": 5}
#         }
#     },
#     "file6": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\22. BAO CAO MMK\DU LIEU\T3\Thong ke thoi gian CD.xlsx",
#         "sheets": {
#             "Thống kê thời gian công đoạn": {"min_cols": 5}
#         }
#     },
#     "file7": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\22. BAO CAO MMK\DU LIEU\T3\2026　VDM管理指標　（　月） QLTB-T3.xlsx",
#         "sheets": {
#             "指標２(外作)": {"min_cols": 5},
#             "指標４（生産性）": {"min_cols": 10},
#         }
#     },
#     "file8": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\22. BAO CAO MMK\DU LIEU\T3\THỜI GIAN LÀM VIỆC VÀ THỜI GIAN TĂNG CA PHÒNG CƠ KHÍ THÁNG 03.2026.xlsx",
#         "sheets": {
#             "GOC": {"min_cols": 8}
#         }
#     },
#     "file9": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\14 .BÂT HỢP CÁCH  2026\1.BHC CÔNG ĐOẠN - KIỂM TRA  2026.xlsx",
#         "sheets": {
#             "BHCKT26": {"min_cols": 1},
#             "BHCCD26": {"min_cols": 20},
#         }
#     },
#     "file10": {
#         "path": r"\\vdm-fsvr\Cokhi-機工\CONG VIEC CHUNG 2026\14 .BÂT HỢP CÁCH  2026\11.KNKH 2026.xlsx",
#         "sheets": {
#             "KHIEU NAI BANG VAN BAN - BC MMK": {"min_cols": 14}
#         }
#     },
# }

# ===== KIỂM TRA FILE / SHEET / CỘT ===== PHỤ ====
FILE_CONFIG = {
    "file1": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3\thoi_gian_gia_cong.xlsx",
        "sheets": {
            "Thời gian gia công": {"min_cols": 12}
        }
    },
    "file2": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3\Download1_2.xlsx",
        "sheets": {
            "Tỉ lệ hoạt động máy": {"min_cols": 3},
            "BV,PCS HT trong tháng theo CĐ": {"min_cols": 3},
            "Nhận BV,PCS trong tháng": {"min_cols": 5},
        }
    },
    "file3": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3\Download2_2026-02-26_2026-03-30.xlsx",
        "sheets": {
            "Số tiền hoàn thành": {"min_cols": 44},
            "Tuân Thủ Kì Hạn": {"min_cols": 18},
        }
    },
    "file4": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3\Download3_2026-02-26 07_00_00_2026-03-30 06_59_00.xlsx",
        "sheets": {
            "BV.PCS Tồn đọng": {"min_cols": 5}
        }
    },
    "file6": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3\Thong ke thoi gian CD.xlsx",
        "sheets": {
            "Thống kê thời gian công đoạn": {"min_cols": 5}
        }
    },
    "file7": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3\2026　VDM管理指標　（　月） QLTB-T3.xlsx",
        "sheets": {
            "指標２(外作)": {"min_cols": 5},
            "指標４（生産性）": {"min_cols": 10},
        }
    },
    "file8": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3\THỜI GIAN LÀM VIỆC VÀ THỜI GIAN TĂNG CA PHÒNG CƠ KHÍ THÁNG 03.2026.xlsx",
        "sheets": {
            "GOC": {"min_cols": 8}
        }
    },
    "file9": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\14 .BAT_HOP_CACH_2026\1.BHC CÔNG ĐOẠN - KIỂM TRA  2026.xlsx",
        "sheets": {
            "BHCKT26": {"min_cols": 1},
            "BHCCD26": {"min_cols": 20},
        }
    },
    "file10": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\14 .BAT_HOP_CACH_2026\KNKH 2026.xlsx",
        "sheets": {
            "KHIEU NAI BANG VAN BAN - BC MMK": {"min_cols": 14}
        }
    },
    "file11": {
        "path": r"D:\Code_cokhi\Bao_Cao_MMK_KPI\T3",
        "sheets": {
            "So_tien_HT": {"min_cols": 14},
        }
    },
}
# ==================================================
# SYSTEM CONFIG
# ==================================================


if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TEMPLATE_PATH = resource_path("2026 VDM KPI.xlsx")
SNO_FIXED_PATH = resource_path("SNO_LIST.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "KET_QUA")


# ===== HẾT =====

# ==================================================
# CONFIG
# ==================================================

FIXED_WORK_DAYS_BY_MONTH = {
    1: 22, 2: 15, 3: 22, 4: 21, 5: 20, 6: 22,
    7: 23, 8: 21, 9: 21, 10: 22, 11: 21, 12: 22
}

KPI_RUN_DAY_BY_MONTH = {
    1: 2, 2: 2, 3: 2, 4: 1, 5: 4, 6: 1,
    7: 1, 8: 3, 9: 3, 10: 1, 11: 2, 12: 1
}

# ==================================================
# FIXED INPUT (KHÔNG UI)
# ==================================================
EXCHANGE_RATE = 26500          # Tỷ giá cố định
KNKH_CASES = 0                 # Không nhập tay

# ===== THÊM HÀM Ở ĐÂY =====
import unicodedata

def normalize_machine_code(value):
    if value is None or pd.isna(value):
        return ""

    # =========================
    # ✅ CHUẨN UNICODE (CỰC KỲ QUAN TRỌNG)
    # EN0１ → EN01
    # ＭＣ０２ → MC02
    # =========================
    s = unicodedata.normalize("NFKC", str(value))

    # --- Chuẩn hóa cơ bản ---
    s = s.upper()

    # Bỏ toàn bộ khoảng trắng
    s = re.sub(r"\s+", "", s)

    # Bỏ các ký tự phân cách thông dụng
    s = re.sub(r"[-_/]", "", s)

    # =========================
    # RULE CỐ ĐỊNH
    # =========================
    if s == "HT":
        return "HT05"

    mc_map = {
        "MC1": "MC01",
        "MC2": "MC02",
        "MC3": "MC03",
    }
    if s in mc_map:
        return mc_map[s]

    # =========================
    # RULE TỔNG QUÁT
    # =========================
    m = re.fullmatch(r"([A-Z]+)0*(\d+)", s)
    if m:
        prefix = m.group(1)
        number = m.group(2)
        if number.isdigit() and len(number) <= 2:
            return f"{prefix}{number.zfill(2)}"
        return f"{prefix}{number}"

    return s
# ===== KẾT THÚC =====

def normalize_cd_from_down1(value):
    """
    Chuẩn hóa CĐ từ Down1:
    MA1, MA2, MA3 -> MA
    LA2 -> LA
    LN2 -> LN
    GR3 -> GR
    AF1, AF2 -> AF
    """
    if value is None or pd.isna(value):
        return ""

    s = str(value).strip().upper()

    # Lấy phần chữ cái đầu (trước số)
    result = ""
    for ch in s:
        if ch.isalpha():
            result += ch
        else:
            break

    return result

def normalize_cd_from_template(value):
    """
    Chuẩn hóa CĐ từ sheet 指標２(内作):
    'LA （旋盤）\\nMÁY TIỆN' -> 'LA'
    """
    if value is None:
        return ""

    s = str(value).strip().upper()

    # Lấy dòng đầu tiên
    s = s.splitlines()[0]

    # Cắt tại khoảng trắng đầu tiên
    s = s.split(" ")[0]

    # Loại ký tự đặc biệt nếu có
    s = s.replace("（", "").replace("(", "")

    return s


def normalize_msyc(value: str) -> str:
    """
    Bỏ toàn bộ ký tự đặc biệt trong MSYC
    VD: L611-01 → L61101
    """
    if not isinstance(value, str):
        return ""
    return re.sub(r"[^A-Z0-9]", "", value.upper())

def build_join_key(syc, rep):
    """
    Tạo JOIN_KEY = SYC + REP
    - SYC: giữ chữ L/R + số, bỏ toàn bộ ký tự đặc biệt
    - REP: chỉ giữ chữ số
    Ví dụ:
        L611-01 , 2  -> L61101-2
        213A/03 , 1 -> 213A03-1
    """
    syc_clean = re.sub(r"[^A-Z0-9]", "", str(syc).strip().upper())
    rep_clean = re.sub(r"[^0-9]", "", str(rep).strip())
    return f"{syc_clean}-{rep_clean}"

def is_valid_code_group(code: str) -> bool:
    """
    Điều kiện mã KHUÔN (MSYC):

    NHÓM 1 (không cần A/B):
        L611, L612, R611, R612, 611, 612

    NHÓM 2 (bắt buộc kết thúc A/B):
        213..., R213... và kết thúc bằng A hoặc B
    """
    if not isinstance(code, str):
        return False

    code = code.strip().upper()

    # ===== NHÓM 1 =====
    group1_prefixes = ("L611", "L612", "R611", "R612", "611", "612")
    for p in group1_prefixes:
        if code.startswith(p):
            return True

    # ===== NHÓM 2 =====
    if (code.startswith("213") or code.startswith("R213")) and code.endswith(("A", "B")):
        return True

    return False

def get_kpi_month_by_monthly_run_day(now=None):
    """
    Mỗi tháng có ngày chốt khác nhau.
    Nếu hôm nay >= ngày chốt của tháng hiện tại:
        → chạy KPI tháng trước
    Nếu chưa tới:
        → chưa chạy

    Trả về (month, year) hoặc None
    """
    if now is None:
        now = datetime.now()

    current_month = now.month
    current_year = now.year
    run_day = KPI_RUN_DAY_BY_MONTH.get(current_month)

    if run_day is None:
        raise Exception(
            f"❌ Chưa cấu hình ngày chốt cho tháng {current_month}"
        )

    if now.day < run_day:
        return None  # chưa đến kỳ chạy

    # chạy KPI tháng trước
    run_month = current_month - 1
    run_year = current_year

    if run_month == 0:
        run_month = 12
        run_year -= 1

    return run_month, run_year

def get_report_month_year(now=None):
    """
    Xác định tháng/năm báo cáo tự động
    Ví dụ:
        Today = 2026-04-13  → Report = 03/2026
        Today = 2026-01-05  → Report = 12/2025
    """
    if now is None:
        now = datetime.now()

    report_month = now.month - 1
    report_year = now.year

    if report_month == 0:
        report_month = 12
        report_year -= 1

    return report_month, report_year

def excel_col(col_index):
    result = ""
    while col_index:
        col_index, rem = divmod(col_index - 1, 26)
        result = chr(65 + rem) + result
    return result

def get_merged_cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return str(cell.value).strip()
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            tl = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return str(tl.value).strip() if tl.value else None
    return None   
            

        #  # FILE 9 - BHC (BHCKT & BHCCD)
        #  # --- BHCKT ---
        # df_bhckt = pd.read_excel(
        #     self.file9_path.get(),
        #     sheet_name="BHCKT26",
        #     header=0
        # )
        # bhckt_month_col = pd.to_numeric(df_bhckt.iloc[:, 0], errors="coerce")
        # bhckt_count = (bhckt_month_col == month).sum()
        # # --- BHCCD ---
        # df_bhccd = pd.read_excel(
        #     self.file9_path.get(),
        #     sheet_name="BHCCD26",
        #     header=0
        # )
        #     bhccd_month_col = pd.to_numeric(df_bhccd.iloc[:, 0], errors="coerce")
        #     bhccd_count = (bhccd_month_col == month).sum()
        #     # =============================
        #     # ĐẾM BHCCD THEO CÔNG ĐOẠN (THEO THÁNG)
        #     # =============================
        #     bhccd_cd_col = df_bhccd.iloc[:, 19].astype(str).str.strip()  # cột T
        #     df_bhccd_m = df_bhccd[bhccd_month_col == month]
        #     bhccd_count_by_cd = (
        #         bhccd_cd_col[bhccd_month_col == month]
        #         .value_counts()
        #         .to_dict()
        #     )
        #     # Ví dụ: {'GS': 2, 'GC': 1}
        #     # =============================
        #     # MAP THÁNG → CỘT
        #     # =============================
        #     start_col_index = 12  # L = tháng 1
        #     target_col_letter = self.get_excel_column_letter(start_col_index + month - 1)
        #     self.log(f"Ghi dữ liệu tháng {month} vào cột {target_col_letter}")
        #     # =============================
        #     # MAP THÁNG → CỘT (指標２(内作))
        #     # =============================
        #     start_col_index_kpi2 = 3  # C = tháng 1
        #     target_col_kpi2 = self.get_excel_column_letter(start_col_index_kpi2 + month - 1)
        #     self.log(f"指標２(内作): ghi tháng {month} vào cột {target_col_kpi2}")
        #     # =============================
        #     # MAP THÁNG → CỘT (指標1)
        #     # =============================
        #     start_col_index_kpi1 = 2  # B = tháng 1
        #     target_col_kpi1 = self.get_excel_column_letter(start_col_index_kpi1 + month - 1)
        #     self.log(f"指標1: ghi tháng {month} vào cột {target_col_kpi1}")
        #     # =============================
        #     # MAP THÁNG → CỘT (指標４)
        #     # Tháng 1 = cột I
        #     # =============================
        #     start_col_index_kpi4 = 9  # I
        #     target_col_kpi4 = self.get_excel_column_letter(start_col_index_kpi4 + month - 1)
        #     # Tháng 1 bắt đầu từ cột F (index = 6)
        #     start_col_index_kpi3 = 6  # F
        #     target_col_kpi3 = self.get_excel_column_letter(start_col_index_kpi3 + month - 1)
        #     self.log(f"指標3: ghi tháng {month} vào cột {target_col_kpi3}")
        #     # =============================
        #     # MAP THÁNG → CỘT (指標２(外作))
        #     # Tháng 1 = cột D
        #     # =============================
        #     start_col_index_kpi2_out = 4  # D
        #     target_col_kpi2_out = self.get_excel_column_letter(start_col_index_kpi2_out + month - 1)
        #     self.log(f"指標２(外作): ghi tháng {month} vào cột {target_col_kpi2_out}")
        #     # =============================
        #     # MAP THÁNG → CỘT (指標４（生産性）)
        #     # Tháng 1 = cột G
        #     # =============================
        #     start_col_index_kpi4_extra = 7  # G
        #     target_col_kpi4_extra = self.get_excel_column_letter(start_col_index_kpi4_extra + month - 1)
        #     self.log(f"指標４（生産性）(extra): ghi tháng {month} vào cột {target_col_kpi4_extra}")
        #     # =============================
        #     # FILE HIỆN TẠI / THÁNG TRƯỚC
        #     # =============================
        #     current_file = os.path.join(OUTPUT_DIR, f"KPI_{year}_{month:02d}.xlsx")
        #     prev_month = month - 1
        #     prev_year = int(year)
        #     if prev_month == 0:
        #         prev_month = 12
        #         prev_year -= 1
        #     prev_file = os.path.join(OUTPUT_DIR, f"KPI_{prev_year}_{prev_month:02d}.xlsx")

        #     # =============================
        #     # LUÔN ĐẢM BẢO KẾ THỪA
        #     # =============================
        #     if not os.path.exists(current_file):
        #         if os.path.exists(prev_file):
        #             copyfile(prev_file, current_file)
        #             self.log(f"Kế thừa dữ liệu từ {os.path.basename(prev_file)}")
        #         else:
        #             copyfile(TEMPLATE_PATH, current_file)
        #             self.log("Tạo file mới từ template")
        #     else:
        #         self.log("File tháng đã tồn tại → chỉ cập nhật thêm dữ liệu")   
        #     # =============================
        #     # ĐỌC DOWN1
        #     # =============================
        #     df = pd.read_excel(
        #         self.file2_path.get(),
        #         sheet_name="Tỉ lệ hoạt động máy",
        #         header=0
        #     )
        #     machine_col = df.iloc[:, 1]   # Cột B
        #     time_col = pd.to_numeric(df.iloc[:, 2], errors="coerce")  # Cột C
        #     machine_time = (
        #         pd.DataFrame({
        #             "machine": machine_col.map(normalize_machine_code),
        #             "time": time_col
        #         })
        #         .groupby("machine")["time"]
        #         .sum()
        #         .to_dict()
        #     )
        #     self.log(f"Tổng hợp {len(machine_time)} máy từ Down1")
        #     # =============================
        #     # ĐỌC DOWN1 - BV,PCS HT trong tháng theo CĐ (CHỈ CỘT A & C)
        #     # =============================
        #     df_bv = pd.read_excel(
        #         self.file2_path.get(),
        #         sheet_name="BV,PCS HT trong tháng theo CĐ",
        #         header=0
        #     )
        #     # Cột A = Công đoạn
        #     # Cột C = Số bản vẽ
        #     cd_col = df_bv.iloc[:, 0].map(normalize_cd_from_down1)   # A
        #     bv_col = pd.to_numeric(df_bv.iloc[:, 2], errors="coerce")  # C
        #     # Tổng số bản vẽ theo CĐ
        #     bv_sum_by_cd = (
        #         pd.DataFrame({
        #             "CD": cd_col.astype(str).str.strip(),
        #             "BV": bv_col
        #         })
        #         .groupby("CD")["BV"]
        #         .sum()
        #         .to_dict()
        #     )
        #     self.log(f"指標２(内作): tổng hợp {len(bv_sum_by_cd)} công đoạn")
        #     # =============================
        #     # ĐỌC DOWN1 - Nhận BV,PCS trong tháng 
        #     # =============================
        #     df_nhan = pd.read_excel(
        #         self.file2_path.get(),
        #         sheet_name="Nhận BV,PCS trong tháng",
        #         header=0
        #     )
        #     # Cột A = STT (chỉ dùng để đếm dòng)
        #     # Cột E = Số lượng
        #     col_stt = df_nhan.iloc[:, 0]
        #     col_qty = pd.to_numeric(df_nhan.iloc[:, 4], errors="coerce")
        #     # ✅ SỐ DÒNG (bỏ header, bỏ dòng trống)
        #     total_rows = col_stt.notna().sum()
        #     # ✅ TỔNG SỐ LƯỢNG
        #     total_quantity = col_qty.sum()
        #     self.log(
        #         f"指標1: Số dòng = {total_rows}, Tổng số lượng = {total_quantity}"
        #     )
        #     # =============================
        #     # ĐỌC DOWN3 - BV.PCS Tồn đọng
        #     # =============================
        #     df_ton = pd.read_excel(
        #         self.file4_path.get(),
        #         sheet_name="BV.PCS Tồn đọng",
        #         header=0
        #     )
        #     # Cột A = STT (đếm dòng)
        #     # Cột E = Số lượng tồn
        #     col_stt_ton = df_ton.iloc[:, 0]
        #     col_qty_ton = pd.to_numeric(df_ton.iloc[:, 4], errors="coerce")
        #     # ✅ SỐ DÒNG (bỏ header, bỏ dòng trống)
        #     total_rows_ton = col_stt_ton.notna().sum()
        #     # ✅ TỔNG SỐ LƯỢNG
        #     total_qty_ton = col_qty_ton.sum()
        #     self.log(
        #         f"指標1 (Tồn đọng): Số dòng = {total_rows_ton}, Tổng số lượng = {total_qty_ton}"
        #     )
        #     # =============================
        #     # ĐỌC DOWN2 - Số tiền hoàn thành (CHỈ 1 LẦN)
        #     # =============================
        #     df_ht = pd.read_excel(
        #         self.file3_path.get(),
        #         sheet_name="Số tiền hoàn thành",
        #         header=0
        #     )
        #     # --- CỘT DỮ LIỆU ---
        #     col_key = df_ht.iloc[:, 0].astype(str).str.strip()   # KEY (cột A)
        #     col_c   = df_ht.iloc[:, 2].astype(str).str.strip()   # Cột C
        #     col_qty = pd.to_numeric(df_ht.iloc[:, 5], errors="coerce")   # Cột F
        #     col_aq  = pd.to_numeric(df_ht.iloc[:, 42], errors="coerce")  # Cột AQ (tiền)
        #     col_ar  = df_ht.iloc[:, 43].astype(str).str.strip()          # Cột AR
        #     # --- ĐIỀU KIỆN LỌC ---
        #     mask_keep = col_ar == "Bản vẽ hoàn thành"
        #     mask_exclude = (
        #         col_key.str.startswith(("412", "L412", "R412")) &
        #         col_c.str.startswith("C")
        #     )
        #     df_valid = df_ht[mask_keep & ~mask_exclude]
        #     df_st = df_ht[mask_keep & ~mask_exclude].copy()
        #     # ==================================================
        #     # ✅ KHUÔN – TÍNH TRỰC TIẾP TỪ TUÂN THỦ KÌ HẠN
        #     # ==================================================
            
        #                 # =============================
        #     # ĐỌC TUÂN THỦ KÌ HẠN
        #     # =============================
        #     df_ttkh = pd.read_excel(
        #         self.file3_path.get(),
        #         sheet_name="Tuân Thủ Kì Hạn",
        #         header=0
        #     )
            
        #     # --- TÍNH TOÁN TỪ Down2 ---
        #     total_rows_ht  = df_valid.iloc[:, 0].notna().sum()
        #     total_qty_ht   = col_qty[mask_keep & ~mask_exclude].sum()
        #     total_money_ht = col_aq[mask_keep & ~mask_exclude].sum()
        #     total_money_usd = total_money_ht / rate
        #     self.log(
        #         f"HT: Số dòng={total_rows_ht}, "
        #         f"Số lượng={total_qty_ht}, "
        #         f"Tổng tiền={total_money_ht}"
        #     )
        #     # =============================
        #     # ĐỌC FILE 6 - Thống kê thời gian công đoạn
        #     # =============================
        #     df_cd_time = pd.read_excel(
        #         self.file6_path.get(),
        #         sheet_name="Thống kê thời gian công đoạn",
        #         header=0
        #     )
        #     # Cột B = mã công đoạn
        #     # Cột E = giá trị cần ghi
        #     cd_key = df_cd_time.iloc[:, 1].astype(str).str.strip()
        #     cd_value = pd.to_numeric(df_cd_time.iloc[:, 4], errors="coerce")
        #     cd_time_map = dict(zip(cd_key, cd_value))
        #     self.log(f"指標3: đọc {len(cd_time_map)} công đoạn từ file thống kê")
        #     # =============================
        #     # ĐỌC FILE 7 - 指標２(外作)
        #     # =============================
        #     df_out = pd.read_excel(
        #         self.file7_path.get(),
        #         sheet_name="指標２(外作)",
        #         header=8
        #     )
        #     df_out.columns = df_out.columns.astype(str).str.strip()
        #     # ✅ FIX MERGE CELL
        #     df_out.iloc[:, 0] = df_out.iloc[:, 0].ffill()
        #     df_out.iloc[:, 1] = df_out.iloc[:, 1].ffill()
        #     # =============================
        #     # 🔥 TÌM CỘT THEO THÁNG
        #     # =============================
        #     month_label = f"{month}月"   # ví dụ : "3月"
        #     target_col_idx = None
        #     for i, col in enumerate(df_out.columns):
        #         if month_label in str(col):
        #             target_col_idx = i
        #             break
        #     if target_col_idx is None:
        #         raise Exception(f"Không tìm thấy cột tháng {month_label} trong file 7")
        #     self.log(f"File 7: lấy dữ liệu cột {month_label}")
        #     # =============================
        #     # ĐỌC FILE 7 - 指標４（生産性）
        #     # =============================
        #     df_kpi4_file7 = pd.read_excel(
        #         self.file7_path.get(),
        #         sheet_name="指標４（生産性）",
        #         header=None
        #     )
        #     # Lấy giá trị
        #     col_index_file7 = 6 + (month - 1)  # G = 6
        #     val_row28 = pd.to_numeric(df_kpi4_file7.iloc[27, col_index_file7], errors="coerce")
        #     val_row29 = pd.to_numeric(df_kpi4_file7.iloc[28, col_index_file7], errors="coerce")
        #     val_row30 = pd.to_numeric(df_kpi4_file7.iloc[29, col_index_file7], errors="coerce")
        #     self.log(f"File7 KPI4: Row29={val_row29}, Row30={val_row30}")
        #     # =============================
        #     # LẤY DỮ LIỆU
        #     # =============================
        #     col_a = df_out.iloc[:, 0].astype(str).str.strip()
        #     col_b = df_out.iloc[:, 1].astype(str).str.strip()
        #     col_c = df_out.iloc[:, 2].astype(str).str.strip()
        #     col_val = pd.to_numeric(df_out.iloc[:, target_col_idx], errors="coerce")
        #     # =============================
        #     # BUILD MAP
        #     # =============================
        #     out_map = {}
        #     for i in range(len(df_out)):
        #         key = (col_a.iloc[i], col_b.iloc[i])
        #         label = col_c.iloc[i]
        #         val = col_val.iloc[i]
        #         if key not in out_map:
        #             out_map[key] = {"bv": 0, "ng": 0}
        #         if label == "図面総数":
        #             out_map[key]["bv"] = val
        #         elif label == "不良件数":
        #             out_map[key]["ng"] = val
        #     out_map = {k: (v["bv"], v["ng"]) for k, v in out_map.items()}
        #     self.log(f"指標２(外作): xử lý {len(out_map)} nhóm")
        #     # =============================
        #     # LẤY KEY (A, B) TỪ SỐ TIỀN HOÀN THÀNH (ĐÃ LỌC)
        #     # =============================
        #     valid_keys = set(
        #         zip(
        #             df_valid.iloc[:, 0].astype(str).str.strip(),  # Cột A
        #             df_valid.iloc[:, 1].astype(str).str.strip()   # Cột B
        #         )
        #     )
        #     tt_a = df_ttkh.iloc[:, 0].astype(str).str.strip()   # Cột A
        #     tt_b = df_ttkh.iloc[:, 1].astype(str).str.strip()   # Cột B
        #     col_o = pd.to_numeric(df_ttkh.iloc[:, 14], errors="coerce")  # Cột O
        #     col_r = pd.to_numeric(df_ttkh.iloc[:, 17], errors="coerce")  # Cột R
        #     # Match theo (A, B)
        #     mask_match = pd.Series(
        #         [(a, b) in valid_keys for a, b in zip(tt_a, tt_b)],
        #         index=df_ttkh.index
        #     )
        #     # =============================
        #     # LỌC THÊM: CỘT A BẮT ĐẦU = "L213"
        #     # (TRÊN DỮ LIỆU ĐÃ MATCH)
        #     # =============================
        #     mask_L213 = (
        #         mask_match &
        #         df_ttkh.iloc[:, 0].astype(str).str.startswith("L213")
        #     )
        #     tt_key = df_ttkh.iloc[:, 0].astype(str).str.strip()      # KEY (cột A)
        #     time_j = pd.to_numeric(df_ttkh.iloc[:, 9], errors="coerce")   # Cột J
        #     time_k = pd.to_numeric(df_ttkh.iloc[:,10], errors="coerce")  # Cột K
        #     total_time_j = time_j[mask_match].sum()/60
        #     total_time_k = time_k[mask_match].sum()/60
        #     self.log(
        #         f"Tuân thủ: Tổng J={total_time_j}, Tổng K={total_time_k}"
        #     )
        #     # =============================
        #     # TÍNH SỐ BẢN VẼ TRỄ HẸN (TUÂN THỦ KÌ HẠN)
        #     # Điều kiện: BV đã hoàn thành + (O - Q) > 0
        #     # =============================
        #     col_o = pd.to_numeric(df_ttkh.iloc[:, 14], errors="coerce")  # Cột O
        #     col_q = pd.to_numeric(df_ttkh.iloc[:, 16], errors="coerce")  # Cột Q
        #     # Trễ hẹn nếu (O - Q) > 0
        #     mask_late = (col_o - col_q) > 0
        #     # ✅ SỐ BV TRỄ HẸN (CHỈ TRONG TẬP ĐÃ HOÀN THÀNH)
        #     late_bv_count = (mask_match & mask_late).sum()
        #     self.log(
        #         f"Trễ hẹn: BV hoàn thành={total_rows_ht}, BV trễ hẹn={late_bv_count}"
        #     )
        #     # =============================
        #     # FILE 8 - GIỜ NHÂN SỰ (GOC)
        #     # CHỈ TÍNH NV CÒN LÀM VIỆC
        #     # =============================
        #     fixed_leave_codes = {
        #         "10064", "10135", "10197", "10202", "10297",
        #         "10444", "10447", "10597", "10649", "10713",
        #         "11139", "11247", "12329", "12443",
        #         "20132", "20839", "49543"
        #     }
        #     df_hr = pd.read_excel(
        #         self.file8_path.get(),
        #         sheet_name="GOC",
        #         header=0
        #     )
        #     # ✅ CHUẨN HÓA MÃ NV – CHỈ GIỮ CHỮ SỐ
        #     mn_col = df_hr.iloc[:, 1].apply(
        #         lambda x: ''.join(filter(str.isdigit, str(x)))
        #     )
            
        #     # ✅ LOẠI DÒNG KHÔNG CÓ MÃ NV
        #     mask_has_nv = mn_col != ""
        #     mn_col = mn_col[mask_has_nv]
        #     df_hr = df_hr.loc[mask_has_nv]
        #     # ===== CỘT K – TRẠNG THÁI HỢP ĐỒNG =====
        #     # index cột K = 10
        #     status_col = df_hr.iloc[:, 10].astype(str).str.upper()
        #     # ✅ NV nghỉ do KTHD
        #     mask_kthd = status_col.str.contains("KTHD", na=False)
        #     # ✅ NV nghỉ do danh sách cố định
        #     mask_fixed_leave = mn_col.isin(fixed_leave_codes)
        #     # ✅ MASK NGHỈ VIỆC CUỐI
        #     mask_leave_final = mask_kthd | mask_fixed_leave
        #     # ✅ MASK NV CÒN LÀM
        #     mask_active_nv = ~mask_leave_final
        #     # ====== ✅ TỔNG SỐ NGƯỜI ======
        #     total_people = mn_col.nunique()
        #     # ====== ✅ NV GIA CÔNG ======
        #     processing_staff = (
        #         mn_col[~mn_col.isin(fixed_leave_codes)].nunique()
        #     )
        #     self.log(
        #         f"File 8: Tổng số người = {total_people}, "
        #         f"NV Gia công = {processing_staff} "
        #         f"(loại fixed + KTHD)"
        #     )
        #     col_g = pd.to_numeric(df_hr.iloc[:, 6], errors="coerce")
        #     col_h = pd.to_numeric(df_hr.iloc[:, 7], errors="coerce")
        #     sum_gh_each = col_g.add(col_h, fill_value=0)
        #     # ✅ GIỜ FILE 8 RIÊNG CHO ROW 18
        #     hours_11928_file8 = sum_gh_each[mn_col == "11928"].sum()
        #     hours_12157_file8 = sum_gh_each[mn_col == "12157"].sum()
        #     # ✅ ĐIỀU KIỆN CHUẨN: PHẢI CÓ MÃ NV + KHÔNG NGHỈ
        #     mask_active_nv_final = (mn_col != "") & (~mn_col.isin(fixed_leave_codes))
        #     total_gh_after_leave = sum_gh_each[mask_active_nv_final].sum()
        #     # ✅ CHỈ TÍNH NV CÒN LÀM VIỆC
        #     total_gh_after_leave = sum_gh_each[mask_active_nv].sum()
        #     self.log(
        #         f"File Giờ NS: Tổng giờ NV trực tiếp (row13) = {total_gh_after_leave:.2f}"
        #     )
        #     # =============================
        #     # FILE 1 - THỜI GIAN GIA CÔNG
        #     # =============================
        #     df_tg = pd.read_excel(
        #         self.file1_path.get(),
        #         sheet_name="Thời gian gia công",
        #         header=0
        #     )

        #     nv_col = df_tg.iloc[:, 7].astype(str).str.strip()   # cột H (mã NV)
        #     type_col = df_tg.iloc[:, 8].astype(str).str.strip() # cột I
        #     time_l = pd.to_numeric(df_tg.iloc[:, 11], errors="coerce")  # cột L (phút)
        #     # NV 11928
        #     total_hours_11928 = time_l[nv_col == "11928"].sum() / 60
        #     final_11928 = total_hours_11928 - hours_11928_file8
        #     # NV 12157
        #     total_hours_12157 = time_l[nv_col == "12157"].sum() / 60
        #     final_12157 = total_hours_12157 - hours_12157_file8
        #     # Loại M
        #     total_hours_M = time_l[type_col == "M"].sum() / 60
        #     self.log(
        #         f"FILE1: 11928={final_11928:.2f}, "
        #         f"12157={final_12157:.2f}, M={total_hours_M:.2f}"
        #     )
        #     tt_code = df_ttkh.iloc[:, 0].astype(str).str.strip().str.upper()
        #     mask_code_valid = tt_code.apply(is_valid_code_group)
        #     # ==================================================
        #     # ✅ KHUÔN – TÍNH ĐÚNG HÀNG 49 & 66 (THEO DÒNG TT)
        #     # ==================================================
        #     # --- LOAD SNO ---
        #     df_sno = pd.read_excel(SNO_FIXED_PATH, header=0)
        #     col_sno = df_sno.iloc[:, 0].astype(str).str.strip().str.upper()
        #     sno_set = set(col_sno)
        #     # --- TUÂN THỦ ---
        #     tt_msyc = df_ttkh.iloc[:, 0].astype(str).str.strip().str.upper()
        #     tt_sno  = df_ttkh.iloc[:, 2].astype(str).str.strip().str.upper()
        #     col_o = pd.to_numeric(df_ttkh.iloc[:, 14], errors="coerce")  # O
        #     col_r = pd.to_numeric(df_ttkh.iloc[:, 17], errors="coerce")  # R
        #     # --- MASK ---
        #     mask_khuon_code = tt_msyc.apply(is_valid_code_group)
        #     mask_sno = tt_sno.isin(sno_set)
        #     mask_khuon_final = mask_match & mask_khuon_code & mask_sno
        #     # --- HÀNG 49 ---
        #     valid_o = col_o[mask_khuon_final]
        #     result_o_khuon = valid_o.mean() if not valid_o.empty else 0
        #     # --- HÀNG 66 ---
        #     valid_r = col_r[mask_khuon_final]
        #     result_r_khuon = valid_r.mean() if not valid_r.empty else 0
        #     # ==================================================
        #     # ✅ MASK KHÁC = MATCH − MÁY TĐ − KHUÔN
        #     # ==================================================
        #     mask_khac = (
        #         mask_match
        #         & ~mask_L213          # loại Máy TĐ
        #         & ~mask_khuon_final  # loại Khuôn
        #     )
        #     # =============================
        #     # LT Tr/B Nhận – H/Thành (Khác)
        #     # =============================
        #     valid_o_khac = col_o[mask_khac]
        #     result_o_khac = valid_o_khac.mean() if not valid_o_khac.empty else 0
        #     self.log(
        #         f"KHÁC - o: COUNT={valid_o_khac.count()}, "
        #         f"MEAN={result_o_khac:.2f}"
        #     )
        #     # =============================
        #     # LT Tr/B G/CÔNG – H/THÀNH (KHÁC)
        #     # =============================
        #     valid_r_khac = col_r[mask_khac]
        #     result_r_khac = valid_r_khac.mean() if not valid_r_khac.empty else 0
        #     self.log(
        #         f"KHÁC - R: COUNT={valid_r_khac.count()}, "
        #         f"MEAN={result_r_khac:.2f}"
        #     )
        #     # ✅ TỬ SỐ: tổng O của các dòng match
        #     sum_o_match = col_o[mask_match].sum()
        #     sum_o_L213 = col_o[mask_L213].sum()
        #     # sum_o_khuon = pd.to_numeric(df_khuon.iloc[:, 14], errors="coerce").sum()
        #     # ✅ TỬ SỐ: tổng R của các dòng match
        #     sum_r_match = col_r[mask_match].sum()
        #     sum_r_L213 = col_r[mask_L213].sum()
        #     # sum_r_khuon = pd.to_numeric(df_khuon.iloc[:, 17], errors="coerce").sum()
        #     # ✅ MẪU SỐ: số dòng có dữ liệu ở cột O (trừ header)
        #     # ✅ CHỈ ĐẾM O Ở NHỮNG DÒNG MATCH
        #     count_o_total = col_o[mask_match].notna().sum()
        #     count_o_L213 = col_o[mask_L213].notna().sum()
        #     # ✅ MẪU SỐ: số dòng có dữ liệu ở cột R (trừ header)
        #     # ✅ CHỈ ĐẾM R Ở NHỮNG DÒNG MATCH
        #     count_r_total = col_r[mask_match].notna().sum()
        #     count_r_L213 = col_r[mask_L213].notna().sum()

        #     # ✅ KẾT QUẢ CUỐI
        #     result_ratio_o = sum_o_match / count_o_total if count_o_total != 0 else 0
        #     result_o_L213 = sum_o_L213 / count_o_L213 if count_o_L213 != 0 else 0
        #     # result_o_khuon = pd.to_numeric(df_khuon.iloc[:, 14], errors="coerce").mean()
        #     result_ratio_r = sum_r_match / count_r_total if count_r_total != 0 else 0
        #     result_r_L213 = sum_r_L213 / count_r_L213 if count_r_L213 != 0 else 0
        #     # result_r_khuon = pd.to_numeric(df_khuon.iloc[:, 17], errors="coerce").mean()
        #     self.log(
        #         f"指標1 row44: SUM(O match)={sum_o_match}, "
        #         f"COUNT(O)={count_o_total}, RESULT={result_ratio_o}"
        #     )
        #     self.log(
        #         f"L213 - O: SUM={sum_o_L213}, COUNT={count_o_L213}, RESULT={result_o_L213}"
        #     )
        #     # self.log(
        #     #     f"Khuon - O: SUM={sum_o_khuon}, COUNT={count_khuon}, RESULT={result_o_khuon}"
        #     # )
            
        #     self.log(
        #         f"指標1 row61: SUM(R match)={sum_r_match}, "
        #         f"COUNT(R match)={count_r_total}, "
        #         f"RESULT={result_ratio_r}"
        #     )
        #     self.log(
        #         f"L213 - R: SUM={sum_r_L213}, COUNT={count_r_L213}, RESULT={result_r_L213}"
        #     )
            
        #     # =============================
        #     # GHI VÀO TEMPLATE
        #     # =============================
        #     wb = openpyxl.load_workbook(current_file)
        #     ws = wb["機械別実績"]
        #     ws_machine = ws
        #     for row in range(5, ws.max_row + 1):
        #         raw_code = self.get_merged_cell_value(ws, row, 3)
        #         machine_code = normalize_machine_code(raw_code)
        #         if not machine_code:
        #             continue
        #         cell = f"{target_col_letter}{row}"
        #         ws[cell] = machine_time.get(machine_code, 0)
        #         ws[cell].number_format = "#,##0"
        #     # HÀng 2 ← Tổng số ngày làm việc
        #     ws_machine[f"{target_col_letter}2"] = work_days
        #     ws_machine[f"{target_col_letter}2"].number_format = "#,##0"
        #     # =============================
        #     # CHỈ SỐ HÀNG 39 - 指標1
        #     # 100 * (BV hoàn thành - BV trễ hẹn) / BV hoàn thành
        #     # =============================
        #     completed_bv = total_rows_ht
        #     late_bv = late_bv_count
        #     if completed_bv > 0:
        #         ratio_bv_on_time = 100 * (completed_bv - late_bv) / completed_bv
        #     else:
        #         ratio_bv_on_time = 0
            
        #     # Tránh chia 0
        #     if completed_bv > 0:
        #         row79_value = 1 - (bhckt_count / completed_bv)
        #         row84_value = 1 - (bhccd_count / completed_bv)
        #     else:
        #         row79_value = 0
        #         row84_value = 0
        #     # =============================
        #     # GHI VÀO 指標1
        #     # Tháng 1 bắt đầu từ cột B
        #     # =============================
        #     ws_kpi1 = wb["指標1"]
        #     # Hàng 19 ← tổng số bản vẽ nhận
        #     ws_kpi1[f"{target_col_kpi1}19"] = total_rows
        #     ws_kpi1[f"{target_col_kpi1}19"].number_format = "#,##0"
        #     # Hàng 4 ← tổng số pcs nhận
        #     ws_kpi1[f"{target_col_kpi1}4"] = total_quantity
        #     ws_kpi1[f"{target_col_kpi1}4"].number_format = "#,##0"
        #     # Hàng 24 ← số bản vẽ tồn đọng
        #     ws_kpi1[f"{target_col_kpi1}24"] = total_rows_ton
        #     ws_kpi1[f"{target_col_kpi1}24"].number_format = "#,##0"
        #     # Hàng 9 ← tổng số pcs tồn đọng
        #     ws_kpi1[f"{target_col_kpi1}9"] = total_qty_ton
        #     ws_kpi1[f"{target_col_kpi1}9"].number_format = "#,##0"
        #     # Hàng 29 ← tổng số bản vẽ hoàn thành
        #     ws_kpi1[f"{target_col_kpi1}29"] = total_rows_ht
        #     ws_kpi1[f"{target_col_kpi1}29"].number_format = "#,##0"
        #     # Hàng 14 ← tổng số pcs hoàn thành
        #     ws_kpi1[f"{target_col_kpi1}14"] = total_qty_ht
        #     ws_kpi1[f"{target_col_kpi1}14"].number_format = "#,##0"
        #     # Hàng 44 ← Tổng số LT Tr/B Nhận-H/Thành
        #     ws_kpi1[f"{target_col_kpi1}44"] = result_ratio_o
        #     ws_kpi1[f"{target_col_kpi1}44"].number_format = "0.00"
        #     # Hàng 61 ← Tổng số LT Tr/B BĐGC-H/Thành
        #     ws_kpi1[f"{target_col_kpi1}61"] = result_ratio_r
        #     ws_kpi1[f"{target_col_kpi1}61"].number_format = "0.00"
        #     # Hàng 53 ← O / count O (L213)
        #     ws_kpi1[f"{target_col_kpi1}53"] = result_o_L213
        #     ws_kpi1[f"{target_col_kpi1}53"].number_format = "0.00"
        #     # Hàng 70 ← R / count R (L213)
        #     ws_kpi1[f"{target_col_kpi1}70"] = result_r_L213
        #     ws_kpi1[f"{target_col_kpi1}70"].number_format = "0.00"
        #     #Hàng 49 ← O (SNO + nhóm A/B)
        #     ws_kpi1[f"{target_col_kpi1}49"] = result_o_khuon
        #     ws_kpi1[f"{target_col_kpi1}49"].number_format = "0.00"
        #     #Hàng 66 ← R (SNO + nhóm A/B)
        #     ws_kpi1[f"{target_col_kpi1}66"] = result_r_khuon
        #     ws_kpi1[f"{target_col_kpi1}66"].number_format = "0.00"
        #     # Hàng 34 ← từ file 7 hàng 28
        #     ws_kpi1[f"{target_col_kpi1}34"] = val_row28
        #     ws_kpi1[f"{target_col_kpi1}34"].number_format = "#,##0"
        #     # HÀNG 39 ← % BV đúng hạn
        #     ws_kpi1[f"{target_col_kpi1}39"] = ratio_bv_on_time / 100
        #     ws_kpi1[f"{target_col_kpi1}39"].number_format = "0.0%"
        #     self.log(f"指標1 row39 = 100 * ({completed_bv} - {late_bv}) / {completed_bv} = {ratio_bv_on_time:.2f}%")
        #     # HÀNG 89 ← Số kiện KNKH
        #     ws_kpi1[f"{target_col_kpi1}89"] = knkh_cases
        #     ws_kpi1[f"{target_col_kpi1}89"].number_format = "#,##0"
        #     self.log(f"指標1 row89 (Số kiện KNKH) = {knkh_cases}")
        #     # HÀNG 79 ← 1 - BHCKT / BV hoàn thành
        #     ws_kpi1[f"{target_col_kpi1}79"] = row79_value
        #     ws_kpi1[f"{target_col_kpi1}79"].number_format = "0.00%"

        #     self.log(f"指標1 row79 = 1 - {bhckt_count}/{completed_bv} = {row79_value:.2%}")
        #     # HÀNG 84 ← 1 - BHCCD / BV hoàn thành
        #     ws_kpi1[f"{target_col_kpi1}84"] = row84_value
        #     ws_kpi1[f"{target_col_kpi1}84"].number_format = "0.00%"
        #     self.log(f"指標1 row84 = 1 - {bhccd_count}/{completed_bv} = {row84_value:.2%}")
        #     # ✅ HÀNG 57 ← LT Tr/B Nhận-H/Thành (Khác)
        #     ws_kpi1[f"{target_col_kpi1}57"] = result_o_khac
        #     ws_kpi1[f"{target_col_kpi1}57"].number_format = "0.00"
        #     # ✅ HÀNG 74 ← LT Tr/B G/công-H/Thành (Khác)
        #     ws_kpi1[f"{target_col_kpi1}74"] = result_r_khac
        #     ws_kpi1[f"{target_col_kpi1}74"].number_format = "0.00"
        #     # =============================
        #     # GHI VÀO 指標２(内作)
        #     # =============================
        #     ws_kpi2 = wb["指標２(内作)"]
        #     row = 9
        #     while row <= ws_kpi2.max_row:
        #         cd_raw = self.get_merged_cell_value(ws_kpi2, row, 1)
        #         cd = normalize_cd_from_template(cd_raw)
        #         if cd:
        #             ws_kpi2[f"{target_col_kpi2}{row}"] = bv_sum_by_cd.get(cd, 0)
        #             ws_kpi2[f"{target_col_kpi2}{row}"].number_format = "#,##0"
        #         row += 3
        #     # Map: công đoạn (LA, LN, MA...) → row Số phế phẩm
        #     kpi2_phepham_row_map = {}
        #     row = 9  # dòng ALL
        #     while row <= ws_kpi2.max_row:
        #         cd_raw = self.get_merged_cell_value(ws_kpi2, row, 1)
        #         if cd_raw:
        #             cd = normalize_cd_from_template(cd_raw)
        #             if cd:
        #                 # dòng Số phế phẩm = dòng hiện tại + 1
        #                 kpi2_phepham_row_map[cd] = row + 1
        #         row += 3
        #     # =============================
        #     # RESET số phế phẩm = 0 cho tất cả công đoạn 内作
        #     # =============================
        #     for phepham_row in kpi2_phepham_row_map.values():
        #         ws_kpi2[f"{target_col_kpi2}{phepham_row}"] = 0
        #         ws_kpi2[f"{target_col_kpi2}{phepham_row}"].number_format = "#,##0"
        #     # =============================
        #     # GHI THÊM VÀO 指標４（生産性）
        #     # =============================
        #     ws_kpi4 = wb["指標４（生産性) "]
        #     # Hàng 38 ← Tổng tiền hoàn thành
        #     ws_kpi4[f"{target_col_kpi4}38"] = total_money_usd
        #     ws_kpi4[f"{target_col_kpi4}38"].number_format = "#,##0"
        #     # Hàng 17 ← Tổng thời gian cột J
        #     ws_kpi4[f"{target_col_kpi4}17"] = total_time_j
        #     ws_kpi4[f"{target_col_kpi4}17"].number_format = "#,##0.00"
        #     # Hàng 16 ← Tổng thời gian cột K
        #     ws_kpi4[f"{target_col_kpi4}16"] = total_time_k
        #     ws_kpi4[f"{target_col_kpi4}16"].number_format = "#,##0.00"
        #     # Hàng 43 ← từ file 7 hàng 29
        #     ws_kpi4[f"{target_col_kpi4}43"] = val_row29
        #     ws_kpi4[f"{target_col_kpi4}43"].number_format = "#,##0"
        #     # Hàng 44 ← từ file 7 hàng 30
        #     ws_kpi4[f"{target_col_kpi4}44"] = val_row30
        #     ws_kpi4[f"{target_col_kpi4}44"].number_format = "#,##0"
        #     # HÀng 9 ← Tổng số người
        #     ws_kpi4[f"{target_col_kpi4}9"] = total_people
        #     ws_kpi4[f"{target_col_kpi4}9"].number_format = "#,##0"
        #     # HÀng 10 ← NV Gia công
        #     ws_kpi4[f"{target_col_kpi4}10"] = processing_staff
        #     ws_kpi4[f"{target_col_kpi4}10"].number_format = "#,##0"
        #     # HÀNG 18 ← T/G làm việc có người (logic RIÊNG, KHÔNG LIÊN QUAN FILE 8)
        #     total_deduct = final_11928 + final_12157 + total_hours_M
        #     row18_value = total_time_j - total_deduct
        #     ws_kpi4[f"{target_col_kpi4}18"] = row18_value
        #     ws_kpi4[f"{target_col_kpi4}18"].number_format = "#,##0"
        #     self.log(
        #         f"指標４ row18 = {total_time_j:.2f} "
        #         f"- (11928:{final_11928:.2f} + 12157:{final_12157:.2f} + M:{total_hours_M:.2f}) "
        #         f"= {row18_value:.2f}"
        #     )
        #     # HÀNG 13 ← T/G NV trực tiếp (FILE 8 – đã loại NV nghỉ)
        #     ws_kpi4[f"{target_col_kpi4}13"] = total_gh_after_leave
        #     ws_kpi4[f"{target_col_kpi4}13"].number_format = "#,##0.00"
        #     self.log(
        #         f"指標４ row13 (GOC sau trừ NV nghỉ) = {total_gh_after_leave:.2f}"
        #     )
        #     # =============================
        #     # GHI VÀO 指標3（工程毎負荷時間）
        #     # =============================
        #     ws_kpi3 = wb["指標3（工程毎負荷時間)"]
        #     for row in range(5, ws_kpi3.max_row + 1):
        #         cd_in_kpi = ws_kpi3[f"C{row}"].value
        #         if cd_in_kpi is None:
        #             continue
        #         cd_in_kpi = str(cd_in_kpi).strip()
        #         if cd_in_kpi in cd_time_map:
        #             ws_kpi3[f"{target_col_kpi3}{row}"] = cd_time_map[cd_in_kpi]
        #             ws_kpi3[f"{target_col_kpi3}{row}"].number_format = "#,##0"
        #     # =============================
        #     # GHI VÀO 指標２(外作)
        #     # =============================
        #     ws_kpi2_out = wb["指標２(外作)"]
        #     row = 12
        #     while row <= ws_kpi2_out.max_row:
        #         a_val = ws_kpi2_out[f"A{row}"].value
        #         b_val = ws_kpi2_out[f"B{row}"].value
        #         if a_val is None or b_val is None:
        #             row += 1
        #             continue
        #         key = (str(a_val).strip(), str(b_val).strip())
        #         label_draw = str(ws_kpi2_out[f"C{row}"].value).strip()
        #         label_ng   = str(ws_kpi2_out[f"C{row+1}"].value).strip()
        #         if key in out_map:
        #             file7_bv, file7_ng = out_map[key]
        #             if label_draw == "Số bản vẽ" and pd.notna(file7_bv):
        #                 ws_kpi2_out[f"{target_col_kpi2_out}{row}"] = file7_bv
        #             if label_ng == "Số phế phẩm" and pd.notna(file7_ng):
        #                 ws_kpi2_out[f"{target_col_kpi2_out}{row+1}"] = file7_ng
        #         row += 3
        #     others_count = 0
        #     for cd, count in bhccd_count_by_cd.items():
        #         if cd in kpi2_phepham_row_map:
        #             phepham_row = kpi2_phepham_row_map[cd]
        #             ws_kpi2[f"{target_col_kpi2}{phepham_row}"] = count
        #             ws_kpi2[f"{target_col_kpi2}{phepham_row}"].number_format = "#,##0"
        #         else:
        #             others_count += count
        #     ws_kpi2[f"{target_col_kpi2}57"] = others_count
        #     ws_kpi2[f"{target_col_kpi2}57"].number_format = "#,##0"
                
        #     wb.save(current_file)
        #     self.log("✅ Hoàn thành")
        #     self.status_label.config(text="Hoàn thành", fg="green")

        # except Exception as e:
        #     self.log(f"❌ Lỗi: {e}")
        #     self.status_label.config(text="Lỗi", fg="red")
        # finally:
        #     self.progress.stop()
        #     self.btn_process.config(state="normal")


    


        
def run_kpi():
    print("=== START KPI BATCH ===")

    # =============================
    # THÁNG / NĂM BÁO CÁO
    # =============================
    month, year = get_report_month_year()
    work_days = FIXED_WORK_DAYS_BY_MONTH.get(month, 0)
    rate = EXCHANGE_RATE
    knkh_cases = KNKH_CASES

    print(f"📅 KPI tháng {month:02d}/{year}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # =============================
    # FILE HIỆN TẠI / THÁNG TRƯỚC
    # =============================
    current_file = os.path.join(OUTPUT_DIR, f"{year} VDM KPI ({month:02d}).xlsx")

    prev_month = month - 1
    prev_year = year
    if prev_month == 0:
        prev_month = 12
        prev_year -= 1

    prev_file = os.path.join(OUTPUT_DIR, f"KPI_{prev_year}_{prev_month:02d}.xlsx")

    if not os.path.exists(current_file):
        if os.path.exists(prev_file):
            copyfile(prev_file, current_file)
            print(f"📂 Kế thừa từ {os.path.basename(prev_file)}")
        else:
            copyfile(TEMPLATE_PATH, current_file)
            print("📄 Tạo file mới từ template")
    else:
        print("📄 File KPI tháng đã tồn tại → cập nhật tiếp")

    # ==================================================
    # FILE 9 – BHC (BHCKT & BHCCD)
    # ==================================================

    path_file9 = FILE_CONFIG["file9"]["path"]

    # --- BHCKT ---
    df_bhckt = pd.read_excel(
        path_file9,
        sheet_name="BHCKT26",
        header=0
    )
    bhckt_month_col = pd.to_numeric(df_bhckt.iloc[:, 0], errors="coerce")
    bhckt_count = (bhckt_month_col == month).sum()

    # --- BHCCD ---
    df_bhccd = pd.read_excel(
        path_file9,
        sheet_name="BHCCD26",
        header=0
    )
    bhccd_month_col = pd.to_numeric(df_bhccd.iloc[:, 0], errors="coerce")
    bhccd_count = (bhccd_month_col == month).sum()

    # =============================
    # ĐẾM BHCCD THEO CÔNG ĐOẠN
    # =============================
    bhccd_cd_col = df_bhccd.iloc[:, 19].astype(str).str.strip()  # cột T
    bhccd_count_by_cd = (
        bhccd_cd_col[bhccd_month_col == month]
        .value_counts()
        .to_dict()
    )

    print(f"✅ BHCKT = {bhckt_count}, BHCCD = {bhccd_count}")

    # =============================
    # MAP THÁNG → CỘT TEMPLATE
    # =============================
    target_col_letter      = excel_col(12 + month - 1)  # 機械別実績
    target_col_kpi1        = excel_col(2  + month - 1)  # 指標1
    target_col_kpi2        = excel_col(3  + month - 1)  # 指標２(内作)
    target_col_kpi3        = excel_col(6  + month - 1)  # 指標3
    target_col_kpi4        = excel_col(9  + month - 1)  # 指標４
    target_col_kpi2_out    = excel_col(4  + month - 1)  # 指標２(外作)
    target_col_kpi4_extra  = excel_col(7  + month - 1)  # 指標４ extra

    print(f"🧭 Cột tháng {month}: {target_col_letter}")

    # ==================================================
    # GHI VÍ DỤ VÀO TEMPLATE (1 KPI MẪU)
    # ==================================================
    wb = openpyxl.load_workbook(current_file)

    ws_kpi1 = wb["指標1"]

    completed_bv = bhckt_count + bhccd_count if (bhckt_count + bhccd_count) > 0 else 0

    if completed_bv > 0:
        row79_value = 1 - (bhckt_count / completed_bv)
    else:
        row79_value = 0

    ws_kpi1[f"{target_col_kpi1}79"].number_format = "0%"


    print("✅ Ghi KPI FILE 9 thành công")
    print("✅ KPI BATCH DONE")

    # ==================================================
    # FILE 2 – DOWNLOAD1
    # ==================================================
    path_file2 = FILE_CONFIG["file2"]["path"]

    # --------------------------------------------------
    # (1) TỈ LỆ HOẠT ĐỘNG MÁY → 機械別実績
    # --------------------------------------------------
    df_machine = pd.read_excel(
        path_file2,
        sheet_name="Tỉ lệ hoạt động máy",
        header=0
    )

    machine_col = df_machine.iloc[:, 1]   # cột B
    time_col = pd.to_numeric(df_machine.iloc[:, 2], errors="coerce")  # cột C

    machine_time = (
        pd.DataFrame({
            "machine": machine_col.map(normalize_machine_code),
            "time": time_col
        })
        .groupby("machine")["time"]
        .sum()
        .to_dict()
    )

    print(f"✅ FILE 2: Tổng hợp {len(machine_time)} máy")

    wb = openpyxl.load_workbook(current_file)
    ws_machine = wb["機械別実績"]

    col_machine = excel_col(12 + month - 1)  # L = tháng 1

    # --- Cache time GJ ---
    gj_total_time = machine_time.get("GJ", 0)
    gj_half_time = gj_total_time / 2 if gj_total_time else 0

    for row in range(5, ws_machine.max_row + 1):
        raw_code = get_merged_cell_value(ws_machine, row, 3)
        mc = normalize_machine_code(raw_code)

        if not mc:
            continue

        cell = f"{col_machine}{row}"

        # =========================
        # ✅ RULE ĐẶC BIỆT: GJ
        # =========================
        if mc == "GJ02":
            ws_machine[cell] = gj_half_time
            ws_machine[cell].number_format = "#,##0"

        elif mc == "GJ03":
            ws_machine[cell] = gj_half_time
            ws_machine[cell].number_format = "#,##0"

        elif mc == "GJ":
            # ❌ KHÔNG ghi trực tiếp GJ
            ws_machine[cell] = 0

        # =========================
        # ✅ RULE BÌNH THƯỜNG
        # =========================
        else:
            ws_machine[cell] = machine_time.get(mc, 0)
            ws_machine[cell].number_format = "#,##0"

    # Tổng số ngày làm việc
    ws_machine[f"{col_machine}2"] = work_days
    ws_machine[f"{col_machine}2"].number_format = "#,##0"

    print(
        f"🔧 GJ split check: GJ total = {gj_total_time}, "
        f"GJ02 = {gj_half_time}, GJ03 = {gj_half_time}"
    )
    # --------------------------------------------------
    # (2) BV,PCS HT TRONG THÁNG THEO CĐ → 指標２(内作)
    # --------------------------------------------------
    df_bv_cd = pd.read_excel(
        path_file2,
        sheet_name="BV,PCS HT trong tháng theo CĐ",
        header=0
    )

    cd_col = df_bv_cd.iloc[:, 0].map(normalize_cd_from_down1)   # cột A
    bv_col = pd.to_numeric(df_bv_cd.iloc[:, 2], errors="coerce")  # cột C

    bv_sum_by_cd = (
        pd.DataFrame({
            "CD": cd_col.astype(str).str.strip(),
            "BV": bv_col
        })
        .groupby("CD")["BV"]
        .sum()
        .to_dict()
    )

    print(f"✅ FILE 2: BV theo công đoạn = {len(bv_sum_by_cd)}")

    # ==================================================
    # 指標２(内作) – BV & PHẾ PHẨM (VÁ HOÀN CHỈNH CUỐI)
    # ==================================================
    ws_kpi2 = wb["指標２(内作)"]
    col_kpi2 = excel_col(3 + month - 1)

    # --------------------------------------------------
    # 1️⃣ LẤY DANH SÁCH CÔNG ĐOẠN CÓ TRONG SHEET
    # (bỏ dòng ALL, bắt đầu từ dòng 12)
    # --------------------------------------------------
    valid_cd_in_sheet = []

    row = 12
    while row <= ws_kpi2.max_row:
        cd_raw = get_merged_cell_value(ws_kpi2, row, 1)
        cd = normalize_cd_from_template(cd_raw)
        if cd:
            valid_cd_in_sheet.append(cd)
        row += 3

    # --------------------------------------------------
    # 2️⃣ GHI BV + PHẾ PHẨM THEO CÔNG ĐOẠN
    # --------------------------------------------------
    row = 12
    while row <= ws_kpi2.max_row:
        cd_raw = get_merged_cell_value(ws_kpi2, row, 1)
        cd = normalize_cd_from_template(cd_raw)

        if cd:
            # BV theo công đoạn
            ws_kpi2[f"{col_kpi2}{row}"] = bv_sum_by_cd.get(cd, 0)
            ws_kpi2[f"{col_kpi2}{row}"].number_format = "#,##0"

            # Phế phẩm theo công đoạn
            ws_kpi2[f"{col_kpi2}{row + 1}"] = bhccd_count_by_cd.get(cd, 0)
            ws_kpi2[f"{col_kpi2}{row + 1}"].number_format = "#,##0"

        row += 3

    # --------------------------------------------------
    # 3️⃣ TÍNH TỔNG BV (ROW 9 – CHỈ THEO SHEET)
    # --------------------------------------------------
    total_bv_all = sum(
        bv_sum_by_cd.get(cd, 0)
        for cd in valid_cd_in_sheet
    )

    ws_kpi2[f"{col_kpi2}9"] = total_bv_all
    ws_kpi2[f"{col_kpi2}9"].number_format = "#,##0"

    # --------------------------------------------------
    # 4️⃣ PHẾ PHẨM KHÁC → HÀNG 57
    # --------------------------------------------------
    phepham_khac = sum(
        count
        for cd, count in bhccd_count_by_cd.items()
        if cd not in valid_cd_in_sheet
    )

    ws_kpi2[f"{col_kpi2}57"] = phepham_khac
    ws_kpi2[f"{col_kpi2}57"].number_format = "#,##0"

    # --------------------------------------------------
    # 5️⃣ TỔNG PHẾ PHẨM (ROW 10 = TRONG SHEET + KHÁC)
    # --------------------------------------------------
    phepham_trong_sheet = sum(
        bhccd_count_by_cd.get(cd, 0)
        for cd in valid_cd_in_sheet
    )

    total_phepham_all = phepham_trong_sheet + phepham_khac

    ws_kpi2[f"{col_kpi2}10"] = total_phepham_all
    ws_kpi2[f"{col_kpi2}10"].number_format = "#,##0"

    # --------------------------------------------------
    # 6️⃣ LOG KIỂM TRA
    # --------------------------------------------------
    print("✅ 指標２(内作) – CHECK HOÀN CHỈNH")
    print("   Công đoạn trong sheet :", valid_cd_in_sheet)
    print("   Tổng BV (Row 9)       :", total_bv_all)
    print("   Phế phẩm trong sheet :", phepham_trong_sheet)
    print("   Phế phẩm KHÁC (57)   :", phepham_khac)
    print("   Tổng phế phẩm (10)   :", total_phepham_all)



    # --------------------------------------------------
    # (3) NHẬN BV,PCS TRONG THÁNG → 指標1
    # --------------------------------------------------
    df_nhan = pd.read_excel(
        path_file2,
        sheet_name="Nhận BV,PCS trong tháng",
        header=0
    )

    col_stt = df_nhan.iloc[:, 0]
    col_code_c = df_nhan.iloc[:, 2].astype(str).str.strip().str.upper()  # CỘT C
    col_qty = pd.to_numeric(df_nhan.iloc[:, 4], errors="coerce")

    # ✅ Loại các dòng có "DC-EN-" ở cột C
    EXCLUDE_PATTERNS = ["DC-EN-"]

    mask_exclude = col_code_c.str.contains(
        "|".join(EXCLUDE_PATTERNS),
        na=False
    )

    # ✅ CHỈ GIỮ DÒNG HỢP LỆ:
    #   - Có STT
    #   - KHÔNG chứa DC-EN-
    mask_valid = col_stt.notna() & ~mask_exclude

    df_nhan_valid = df_nhan[mask_valid]

    # =============================
    # TÍNH TOÁN SAU LỌC
    # =============================
    total_rows = df_nhan_valid.iloc[:, 0].notna().sum()
    total_quantity = pd.to_numeric(
        df_nhan_valid.iloc[:, 4],
        errors="coerce"
    ).sum()

    print(
        "✅ NHẬN BV,PCS – AFTER FILTER\n"
        f"   Tổng BV hợp lệ = {total_rows}\n"
        f"   Tổng PCS hợp lệ = {total_quantity}\n"
        f"   Bị loại DC-EN- = {mask_exclude.sum()}"
    )


    ws_kpi1 = wb["指標1"]
    col_kpi1 = excel_col(2 + month - 1)  # B = tháng 1

    # Hàng 19: tổng bản vẽ nhận
    ws_kpi1[f"{col_kpi1}19"] = total_rows
    ws_kpi1[f"{col_kpi1}19"].number_format = "#,##0"

    # Hàng 4: tổng PCS nhận
    ws_kpi1[f"{col_kpi1}4"] = total_quantity
    ws_kpi1[f"{col_kpi1}4"].number_format = "#,##0"

    # --------------------------------------------------
    # SAVE SAU FILE 2
    # --------------------------------------------------

    print("✅ FILE 2: Ghi dữ liệu xong")

    # ==================================================
    # FILE 4 – BV.PCS TỒN ĐỌNG
    # ==================================================
    path_file4 = FILE_CONFIG["file4"]["path"]

    df_ton = pd.read_excel(
        path_file4,
        sheet_name="BV.PCS Tồn đọng",
        header=0
    )
    
    col_stt = df_ton.iloc[:, 0]
    col_code_c = df_ton.iloc[:, 2].astype(str).str.strip().str.upper()  # ✅ CỘT C
    col_qty_ton = pd.to_numeric(df_ton.iloc[:, 4], errors="coerce")

    # --------------------------------------------------
    # ✅ PRE-FILTER: LOẠI DC-EN-
    # --------------------------------------------------
    EXCLUDE_PATTERNS = ["DC-EN-"]

    mask_exclude_dc_en = col_code_c.str.contains(
        "|".join(EXCLUDE_PATTERNS),
        na=False
    )

    # --------------------------------------------------
    # ✅ GIỮ DÒNG HỢP LỆ:
    #   - Có STT
    #   - PCS > 0
    #   - KHÔNG phải DC-EN-
    # --------------------------------------------------
    mask_valid = (
        col_stt.notna() &
        (col_qty_ton > 0) &
        ~mask_exclude_dc_en
    )

    df_ton_valid = df_ton[mask_valid] 
    total_rows_ton = df_ton_valid.iloc[:, 0].notna().sum()
    total_qty_ton = pd.to_numeric(
        df_ton_valid.iloc[:, 4],
        errors="coerce"
    ).sum()


    print(
        "✅ FILE 4 – BV.PCS TỒN ĐỌNG (AFTER FILTER)\n"
        f"   BV tồn hợp lệ     = {total_rows_ton}\n"
        f"   PCS tồn hợp lệ    = {total_qty_ton}\n"
        f"   Bị loại DC-EN-    = {mask_exclude_dc_en.sum()}"
    )


    ws_kpi1 = wb["指標1"]
    col_kpi1 = excel_col(2 + month - 1)

    # Row 24: số BV tồn
    ws_kpi1[f"{col_kpi1}24"] = total_rows_ton
    ws_kpi1[f"{col_kpi1}24"].number_format = "#,##0"

    # Row 9: PCS tồn
    ws_kpi1[f"{col_kpi1}9"] = total_qty_ton
    ws_kpi1[f"{col_kpi1}9"].number_format = "#,##0"

    # print(f"✅ FILE 4: BV tồn={total_rows_ton}, PCS tồn={total_qty_ton}")

    # ==================================================
    # FILE 3 – SỐ TIỀN HOÀN THÀNH
    # ==================================================
    path_file3 = FILE_CONFIG["file3"]["path"]

    df_ht = pd.read_excel(
        path_file3,
        sheet_name="Số tiền hoàn thành",
        header=0
    )

    col_key = df_ht.iloc[:, 0].astype(str).str.strip()
    col_c   = df_ht.iloc[:, 2].astype(str).str.strip().str.upper()  # ✅ CỘT C
    col_qty = pd.to_numeric(df_ht.iloc[:, 5], errors="coerce")
    col_aq  = pd.to_numeric(df_ht.iloc[:, 42], errors="coerce")
    col_ar  = df_ht.iloc[:, 43].astype(str).str.strip()

    # =============================
    # ✅ PRE-FILTER: LOẠI DC-EN-
    # =============================
    EXCLUDE_PATTERNS = ["DC-EN-"]

    mask_exclude_dc_en = col_c.str.contains(
        "|".join(EXCLUDE_PATTERNS),
        na=False
    )

    df_ht_pf = df_ht[~mask_exclude_dc_en]

    # =============================
    # ✅ BUSINESS FILTER (GIỮ NGUYÊN LOGIC CŨ)
    # =============================
    col_key_pf = df_ht_pf.iloc[:, 0].astype(str).str.strip()
    col_c_pf   = df_ht_pf.iloc[:, 2].astype(str).str.strip()
    col_ar_pf  = df_ht_pf.iloc[:, 43].astype(str).str.strip()


    mask_keep = col_ar == "Bản vẽ hoàn thành"
    mask_exclude = col_key.str.startswith(("412", "L412", "R412")) & col_c.str.startswith("C")

    df_valid = df_ht[mask_keep & ~mask_exclude]

    total_rows_ht = df_valid.shape[0]
    total_qty_ht = col_qty[mask_keep & ~mask_exclude].sum()
    total_money_vnd = col_aq[mask_keep & ~mask_exclude].sum()
    total_money_usd = total_money_vnd / EXCHANGE_RATE

    ws_kpi1 = wb["指標1"]
    ws_kpi4 = wb["指標４（生産性) "]

    col_kpi1 = excel_col(2 + month - 1)
    col_kpi4 = excel_col(9 + month - 1)

    # 指標1
    ws_kpi1[f"{col_kpi1}29"] = total_rows_ht
    ws_kpi1[f"{col_kpi1}14"] = total_qty_ht

    # # 指標4
    # ws_kpi4[f"{col_kpi4}38"] = total_money_usd
    # ws_kpi4[f"{col_kpi4}38"].number_format = "#,##0"


    print(
        "✅ FILE 3 – SỐ TIỀN HOÀN THÀNH (AFTER FILTER)\n"
        f"   Dòng hợp lệ        = {total_rows_ht}\n"
        f"   Tổng PCS           = {total_qty_ht}\n"
        # f"   Tổng tiền (VND)    = {total_money_vnd:,.0f}\n"
        f"   Bị loại DC-EN-     = {mask_exclude_dc_en.sum()}"
    )

    # ==================================================
    # FILE 3 (TIẾP) – TUÂN THỦ KÌ HẠN
    # TÍNH LT: KHUÔN / KHÁC / L213
    # ==================================================
    df_ttkh = pd.read_excel(
        path_file3,
        sheet_name="Tuân Thủ Kì Hạn",
        header=0
    )
    
    # --------------------------------------------------
    # ✅ PRE-FILTER: LOẠI DC-EN- Ở CỘT C
    # --------------------------------------------------
    col_code_c = df_ttkh.iloc[:, 2].astype(str).str.strip().str.upper()

    EXCLUDE_PATTERNS = ["DC-EN-"]

    mask_exclude_dc_en = col_code_c.str.contains(
        "|".join(EXCLUDE_PATTERNS),
        na=False
    )

    df_ttkh_pf = df_ttkh[~mask_exclude_dc_en]

    print(
        "✅ TUÂN THỦ KÌ HẠN – AFTER DC-EN FILTER\n"
        f"   Tổng dòng ban đầu = {len(df_ttkh)}\n"
        f"   Bị loại DC-EN-    = {mask_exclude_dc_en.sum()}\n"
        f"   Còn lại xử lý     = {len(df_ttkh_pf)}"
    )


    # ---- 1) TẬP KEY ĐÃ HOÀN THÀNH (từ FILE 3 – Số tiền hoàn thành)
    valid_keys = set(
        zip(
            df_valid.iloc[:, 0].astype(str).str.strip(),  # Cột A
            df_valid.iloc[:, 1].astype(str).str.strip()   # Cột B
        )
    )

    tt_a = df_ttkh.iloc[:, 0].astype(str).str.strip()
    tt_b = df_ttkh.iloc[:, 1].astype(str).str.strip()

    # Match theo (A, B)
    mask_match = pd.Series(
        [(a, b) in valid_keys for a, b in zip(tt_a, tt_b)],
        index=df_ttkh.index
    )

    # ---- 2) CỘT LT
    col_o = pd.to_numeric(df_ttkh.iloc[:, 14], errors="coerce")  # LT Ngày nhận (O)
    col_r = pd.to_numeric(df_ttkh.iloc[:, 17], errors="coerce")  # LT sao (R)

    # ---- 3) PHÂN NHÓM
    tt_msyc = df_ttkh.iloc[:, 0].astype(str).str.strip().str.upper()
    tt_sno  = df_ttkh.iloc[:, 2].astype(str).str.strip().str.upper()

    # --- LOAD SNO LIST (đi theo template)
    df_sno = pd.read_excel(SNO_FIXED_PATH, header=0)
    sno_set = set(df_sno.iloc[:, 0].astype(str).str.strip().str.upper())

    # (a) L213
    mask_L213 = mask_match & tt_msyc.str.startswith("L213")

    # (b) KHUÔN: mã hợp lệ + thuộc SNO
    mask_khuon_code = tt_msyc.apply(is_valid_code_group)
    mask_khuon = mask_match & mask_khuon_code & tt_sno.isin(sno_set)

    # (c) KHÁC: match nhưng không phải L213, không phải Khuôn
    mask_khac = mask_match & ~mask_L213 & ~mask_khuon

    # ---- 4) TÍNH TRUNG BÌNH
    def mean_or_zero(series):
        s = series.dropna()
        return float(s.mean()) if not s.empty else 0.0

    # LT Ngày nhận (O)
    lt_khuon_o = mean_or_zero(col_o[mask_khuon])
    lt_L213_o  = mean_or_zero(col_o[mask_L213])
    lt_khac_o  = mean_or_zero(col_o[mask_khac])

    # LT sao (R)
    lt_khuon_r = mean_or_zero(col_r[mask_khuon])
    lt_L213_r  = mean_or_zero(col_r[mask_L213])
    lt_khac_r  = mean_or_zero(col_r[mask_khac])

    # ---- 5) GHI VÀO 指標1
    ws_kpi1 = wb["指標1"]
    col_kpi1 = excel_col(2 + month - 1)

    # Quy ước hàng (theo logic bạn đang dùng):
    # 49: LT Khuôn – O | 66: LT Khuôn – R
    ws_kpi1[f"{col_kpi1}49"] = lt_khuon_o
    ws_kpi1[f"{col_kpi1}49"].number_format = "0.0"
    ws_kpi1[f"{col_kpi1}66"] = lt_khuon_r
    ws_kpi1[f"{col_kpi1}66"].number_format = "0.0"

    # 53: LT L213 – O | 70: LT L213 – R
    ws_kpi1[f"{col_kpi1}53"] = lt_L213_o
    ws_kpi1[f"{col_kpi1}53"].number_format = "0.0"
    ws_kpi1[f"{col_kpi1}70"] = lt_L213_r
    ws_kpi1[f"{col_kpi1}70"].number_format = "0.0"

    # 57: LT Khác – O | 74: LT Khác – R
    ws_kpi1[f"{col_kpi1}57"] = lt_khac_o
    ws_kpi1[f"{col_kpi1}57"].number_format = "0.0"
    ws_kpi1[f"{col_kpi1}74"] = lt_khac_r
    ws_kpi1[f"{col_kpi1}74"].number_format = "0.0"

    print(
        f"✅ LT Khuôn(O/R) = {lt_khuon_o:.2f}/{lt_khuon_r:.2f} | "
        f"L213(O/R) = {lt_L213_o:.2f}/{lt_L213_r:.2f} | "
        f"Khác(O/R) = {lt_khac_o:.2f}/{lt_khac_r:.2f}"
    )
    # ==================================================
    # LT TỔNG – TR/B NHẬN & TR/B BĐGC (TOÀN BỘ)
    # ==================================================
    def mean_or_zero(series):
        s = series.dropna()
        return float(s.mean()) if not s.empty else 0.0

    lt_total_o = mean_or_zero(col_o[mask_match])  # LT Ngày nhận – Tổng
    lt_total_r = mean_or_zero(col_r[mask_match])  # LT sao – Tổng

    ws_kpi1 = wb["指標1"]
    col_kpi1 = excel_col(2 + month - 1)

    # HÀNG 44 – LT Tr/B Nhận – Hoàn thành
    ws_kpi1[f"{col_kpi1}44"] = lt_total_o
    ws_kpi1[f"{col_kpi1}44"].number_format = "0.0"

    # HÀNG 61 – LT Tr/B BĐGC – Hoàn thành
    ws_kpi1[f"{col_kpi1}61"] = lt_total_r
    ws_kpi1[f"{col_kpi1}61"].number_format = "0.0"

    print(
        f"✅ LT TỔNG | Ngày nhận = {lt_total_o:.2f} | "
        f"LT sao = {lt_total_r:.2f}"
    )
    # ==================================================
    # BHC – TỶ LỆ (HÀNG 79 & 84)
    # ==================================================
    if completed_bv > 0:
        row79_value = 1 - (bhckt_count / completed_bv)
        row84_value = 1 - (bhccd_count / completed_bv)
    else:
        row79_value = 0
        row84_value = 0

    ws_kpi1 = wb["指標1"]
    col_kpi1 = excel_col(2 + month - 1)

    # HÀNG 79 – 1 - BHCKT / BV hoàn thành
    ws_kpi1[f"{col_kpi1}79"] = row79_value
    ws_kpi1[f"{col_kpi1}79"].number_format = "0%"

    # HÀNG 84 – CỐ ĐỊNH = 1 (100%)
    ws_kpi1[f"{col_kpi1}84"] = 1
    ws_kpi1[f"{col_kpi1}84"].number_format = "0%"

    print(
        f"✅ 指標1 CHECK\n"
        f"   Row79 = {row79_value:.2%}\n"
        f"   Row84 = 100% (FIXED)"
    )

    # ==================================================
    # 指標1 – HÀNG 34 (từ FILE 7 – row 28)
    # ==================================================
    df_kpi4_file7 = pd.read_excel(
        FILE_CONFIG["file7"]["path"],
        sheet_name="指標４（生産性）",
        header=None
    )

    col_idx_file7 = 6 + (month - 1)  # G = tháng 1
    val_row28 = pd.to_numeric(df_kpi4_file7.iloc[27, col_idx_file7], errors="coerce")

    ws_kpi1 = wb["指標1"]
    col_kpi1 = excel_col(2 + month - 1)

    ws_kpi1[f"{col_kpi1}34"] = val_row28
    ws_kpi1[f"{col_kpi1}34"].number_format = "#,##0"
    # ==================================================
    # 指標1 – HÀNG 39 (% BV đúng hạn)
    # ==================================================
    # Trễ hạn nếu (O - Q) > 0
    col_q = pd.to_numeric(df_ttkh.iloc[:, 16], errors="coerce")  # Cột Q
    mask_late = (col_o - col_q) > 0

    late_bv_count = (mask_match & mask_late).sum()

    if total_rows_ht > 0:
        ratio_bv_on_time = (total_rows_ht - late_bv_count) / total_rows_ht
    else:
        ratio_bv_on_time = 0

    ws_kpi1[f"{col_kpi1}39"] = ratio_bv_on_time
    ws_kpi1[f"{col_kpi1}39"].number_format = "0%"

    # ==================================================
    # FILE 6 – THỐNG KÊ THỜI GIAN CÔNG ĐOẠN
    # ==================================================
    path_file6 = FILE_CONFIG["file6"]["path"]

    df_cd_time = pd.read_excel(
        path_file6,
        sheet_name="Thống kê thời gian công đoạn",
        header=0
    )

    cd_key = df_cd_time.iloc[:, 1].astype(str).str.strip()
    cd_value = pd.to_numeric(df_cd_time.iloc[:, 4], errors="coerce")
    cd_time_map = dict(zip(cd_key, cd_value))

    ws_kpi3 = wb["指標3（工程毎負荷時間)"]
    col_kpi3 = excel_col(6 + month - 1)

    for row in range(5, ws_kpi3.max_row + 1):
        cd = ws_kpi3[f"C{row}"].value
        if cd and cd in cd_time_map:
            ws_kpi3[f"{col_kpi3}{row}"] = cd_time_map[cd]
            ws_kpi3[f"{col_kpi3}{row}"].number_format = "#,##0"

    print("✅ FILE 6: Ghi thời gian công đoạn xong")

    # ==================================================
    # FILE 7 – 指標２(外作)
    # ==================================================
    path_file7 = FILE_CONFIG["file7"]["path"]

    df_out = pd.read_excel(
        path_file7,
        sheet_name="指標２(外作)",
        header=8
    )

    df_out.iloc[:, 0] = df_out.iloc[:, 0].ffill()
    df_out.iloc[:, 1] = df_out.iloc[:, 1].ffill()

    month_label = f"{month}月"
    target_col_idx = next(i for i, c in enumerate(df_out.columns) if month_label in str(c))

    col_a = df_out.iloc[:, 0].astype(str).str.strip()
    col_b = df_out.iloc[:, 1].astype(str).str.strip()
    col_c = df_out.iloc[:, 2].astype(str).str.strip()
    col_val = pd.to_numeric(df_out.iloc[:, target_col_idx], errors="coerce")

    out_map = {}
    for i in range(len(df_out)):
        key = (col_a.iloc[i], col_b.iloc[i])
        if key not in out_map:
            out_map[key] = {"bv": 0, "ng": 0}
        if col_c.iloc[i] == "図面総数":
            out_map[key]["bv"] = col_val.iloc[i]
        elif col_c.iloc[i] == "不良件数":
            out_map[key]["ng"] = col_val.iloc[i]

    ws_kpi2_out = wb["指標２(外作)"]
    col_kpi2_out = excel_col(4 + month - 1)

    row = 12
    while row <= ws_kpi2_out.max_row:
        key = (
            str(ws_kpi2_out[f"A{row}"].value).strip(),
            str(ws_kpi2_out[f"B{row}"].value).strip()
        )
        if key in out_map:
            ws_kpi2_out[f"{col_kpi2_out}{row}"] = out_map[key]["bv"]
            ws_kpi2_out[f"{col_kpi2_out}{row+1}"] = out_map[key]["ng"]
        row += 3

    print("✅ FILE 7: 外作 xong")
    
    # ==================================================
    # FILE 8 – GIỜ NHÂN SỰ (GOC)
    # ==================================================
    fixed_leave_codes = {
         "10064", "10135", "10197", "10202", "10297",
         "10444", "10447", "10597", "10649", "10713",
         "11139", "11247", "12329", "12443",
         "20132", "20839", "49543"
    }
    path_file8 = FILE_CONFIG["file8"]["path"]

    df_hr = pd.read_excel(
        path_file8,
        sheet_name="GOC",
        header=0
    )

    # Cột B = Mã NV
    # ==================================================
    mn_col = df_hr.iloc[:, 1].apply(
        lambda x: ''.join(filter(str.isdigit, str(x)))
    )

    # --------------------------------------------------
    # CỘT A – STT (index 0)
    # --------------------------------------------------
    stt_col = pd.to_numeric(df_hr.iloc[:, 0], errors="coerce")
    # --------------------------------------------------
    # CỘT K – GHI CHÚ (index 10)
    # --------------------------------------------------
    note_col = df_hr.iloc[:, 10].astype(str).str.strip().str.upper()

    # --------------------------------------------------
    # ✅ NGƯỜI HỢP LỆ:
    #   - Có STT
    #   - KHÔNG phải TS
    # --------------------------------------------------
    mask_row9 = stt_col.notna() & (note_col != "TS")
    total_people = mask_row9.sum()

    # Điều kiện loại thêm
    mask_leave_status = note_col.isin(["KTHD", "DCLV"])
    mask_fixed_leave = mn_col.isin(fixed_leave_codes)

    mask_row10 = mask_row9 & ~(mask_leave_status | mask_fixed_leave)
    processing_staff = mask_row10.sum()


    # --------------------------------------------------
    # ✅ CỘT GIỜ (KHÔNG LIÊN QUAN HÀNG 9)
    # --------------------------------------------------
    h1 = pd.to_numeric(df_hr.iloc[:, 6], errors="coerce")  # cột G
    h2 = pd.to_numeric(df_hr.iloc[:, 7], errors="coerce")  # cột H
    total_hours = (h1 + h2)[mask_row10].sum()


    ws_kpi4 = wb["指標４（生産性) "]
    col_kpi4 = excel_col(9 + month - 1)

    
    # ✅ HÀNG 9 – TỔNG SỐ NGƯỜI (ĐÃ LOẠI TS)
    ws_kpi4[f"{col_kpi4}9"]  = total_people
    ws_kpi4[f"{col_kpi4}9"].number_format = "#,##0"

     # ✅ HÀNG 10 – TỔNG SỐ NGƯỜI TT  (ĐÃ LOẠI TS-KTHD)
    ws_kpi4[f"{col_kpi4}10"] = processing_staff
    ws_kpi4[f"{col_kpi4}10"].number_format = "#,##0"

    # ✅ HÀNG 13 – TỔNG GIỜ (GIỮ NGUYÊN)
    ws_kpi4[f"{col_kpi4}13"] = total_hours
    ws_kpi4[f"{col_kpi4}13"].number_format = "#,##0"

    print(
        f"✅ FILE 8 – GOC (CHECK)\n"
        f"   Hàng 9 – Tổng người (loại TS)        = {total_people}\n"
        f"   Hàng 10 – NV gia công (loại KTHD + fixed, không trừ đúp) = {processing_staff}\n"
        f"   Trong đó:\n"
        f"     • KTHD        = {mask_leave_status.sum()}\n"
        f"     • Fixed leave = {mask_fixed_leave.sum()}"
        f"   Hàng 13 – Tổng giờ NV GC    = {total_hours:.2f}"
    )

    val_row29 = pd.to_numeric(df_kpi4_file7.iloc[28, col_idx_file7], errors="coerce")
    val_row30 = pd.to_numeric(df_kpi4_file7.iloc[29, col_idx_file7], errors="coerce")

    ws_kpi4[f"{col_kpi4}43"] = val_row29
    ws_kpi4[f"{col_kpi4}43"].number_format = "#,##0"

    ws_kpi4[f"{col_kpi4}44"] = val_row30
    ws_kpi4[f"{col_kpi4}44"].number_format = "#,##0"

    sum_gh_each = h1.add(h2, fill_value=0)
    # ===== VÁ: đọc FILE 1 – Thời gian gia công (cho HÀNG 18) =====
    df_tg = pd.read_excel(
        FILE_CONFIG["file1"]["path"],
        sheet_name="Thời gian gia công",
        header=0
    )

    # ==================================================
    # 指標４（生産性) – HÀNG 16 & 17 (THEO COMMENT GỐC)
    # J, K là PHÚT → đổi sang GIỜ
    # ==================================================

    # Cột J, K trong sheet "Tuân Thủ Kì Hạn"
    time_j = pd.to_numeric(df_ttkh.iloc[:, 9], errors="coerce")    # J (phút)
    time_k = pd.to_numeric(df_ttkh.iloc[:, 10], errors="coerce")  # K (phút)

    total_time_j_hours = time_j[mask_match].sum() / 60
    total_time_k_hours = time_k[mask_match].sum() / 60

    ws_kpi4 = wb["指標４（生産性) "]
    col_kpi4 = excel_col(9 + month - 1)

    # HÀNG 17 – Tổng thời gian J (GIỜ)
    ws_kpi4[f"{col_kpi4}17"] = total_time_j_hours
    ws_kpi4[f"{col_kpi4}17"].number_format = "#,##0"

    # HÀNG 16 – Tổng thời gian K (GIỜ)
    ws_kpi4[f"{col_kpi4}16"] = total_time_k_hours
    ws_kpi4[f"{col_kpi4}16"].number_format = "#,##0"

    print(
        f"🕒 指標４（生産性) | Tháng {month:02d}/{year} | Cột {col_kpi4}\n"
        f"   • Hàng 16 (Tổng giờ K): {total_time_k_hours:.2f}\n"
        f"   • Hàng 17 (Tổng giờ J): {total_time_j_hours:.2f}\n"
    )

    # ==================================================
    # 指標４（生産性) – HÀNG 18 (VÁ + LOG CHI TIẾT)
    # ==================================================

    print("\n================ ROW 18 – BEGIN DEBUG ================")

    # ==================================================
    # FILE 8 – GIỜ NHÂN SỰ (GOC) – CHỈ NV CÒN LÀM VIỆC
    # ==================================================
    hours_11928_file8 = sum_gh_each[(mn_col == "11928") & mask_row10].sum()
    hours_12157_file8 = sum_gh_each[(mn_col == "12157") & mask_row10].sum()

    print(
        "📘 FILE 8 – GOC (GIỜ CÔNG)\n"
        f"   NV 11928 = {hours_11928_file8:.2f} giờ\n"
        f"   NV 12157 = {hours_12157_file8:.2f} giờ"
    )

    # ==================================================
    # FILE 1 – THỜI GIAN GIA CÔNG
    # ==================================================
    machine_col = df_tg.iloc[:, 6].astype(str).str.strip()   # CỘT G – SỐ MÁY
    nv_col      = df_tg.iloc[:, 7].astype(str).str.strip()
    type_col    = df_tg.iloc[:, 8].astype(str).str.strip()
    time_l      = pd.to_numeric(df_tg.iloc[:, 11], errors="coerce")  # phút

    # --------------------------------------------------
    # FILE 1 – GIỜ NV 11928
    # --------------------------------------------------
    total_hours_11928 = time_l[nv_col == "11928"].sum() / 60

    # --------------------------------------------------
    # FILE 1 – GIỜ MÁY AA (KHÔNG TÍNH PHIẾU M)
    # --------------------------------------------------
    mask_AA_no_M = (machine_col == "AA") & (type_col != "M")
    total_hours_AA_no_M = time_l[mask_AA_no_M].sum() / 60

    # --------------------------------------------------
    # FILE 1 – GIỜ LOẠI M (KHÔNG TRÙNG NV 11928)
    # --------------------------------------------------
    mask_M_only = (type_col == "M") & (~nv_col.isin(["11928"]))
    total_hours_M = time_l[mask_M_only].sum() / 60

    # ==================================================
    # NET SAU TRỪ FILE 8 (KHÓA ÂM)
    # ==================================================
    final_11928 = max(0, total_hours_11928 - hours_11928_file8)
    final_12157 = max(0, total_hours_AA_no_M - hours_12157_file8)

    print(
        "➖ NET SAU TRỪ FILE 8\n"
        f"   NV 11928  = {total_hours_11928:.2f} - {hours_11928_file8:.2f}"
        f" → {final_11928:.2f}\n"
        f"   NV 12157* = {total_hours_AA_no_M:.2f} - {hours_12157_file8:.2f}"
        f" → {final_12157:.2f}  (MÁY AA – NO M)"
    )

    # ==================================================
    # TÍNH ROW 18
    # ==================================================
    total_deduct = final_11928 + final_12157 + total_hours_M
    row18_value = max(0, total_time_j_hours - total_deduct)

    print(
        "🧮 TÍNH HÀNG 18\n"
        f"   Tổng J (giờ)   = {total_time_j_hours:.2f}\n"
        f"   Trừ 11928     = {final_11928:.2f}\n"
        f"   Trừ AA (12157)= {final_12157:.2f}\n"
        f"   Trừ M         = {total_hours_M:.2f}\n"
        f"   → ROW 18 NET  = {row18_value:.2f}"
    )

    # ==================================================
    # GHI EXCEL
    # ==================================================
    ws_kpi4[f"{col_kpi4}18"] = row18_value
    ws_kpi4[f"{col_kpi4}18"].number_format = "#,##0"

    print("================= ROW 18 – END DEBUG =================\n")



    # ==================================================
    # FILE 10 – KHIẾU NẠI BẰNG VĂN BẢN (THEO THÁNG)
    # ==================================================
    path_file10 = FILE_CONFIG["file10"]["path"]

    df_kn = pd.read_excel(
        path_file10,
        sheet_name="KHIEU NAI BANG VAN BAN - BC MMK",
        header=None
    )

    # -----------------------------
    # DÒNG "Số kiện phát sinh" = row 3 (index 2)
    # CỘT THÁNG: T1 = C (index 2)
    # -----------------------------
    row_so_kien = 2
    col_thang = 2 + (month - 1)

    try:
        knkh_cases = int(df_kn.iloc[row_so_kien, col_thang])
    except Exception:
        knkh_cases = 0

    # GHI VÀO 指標1 – HÀNG 89
    ws_kpi1 = wb["指標1"]
    col_kpi1 = excel_col(2 + month - 1)  # B = tháng 1

    ws_kpi1[f"{col_kpi1}89"] = knkh_cases
    ws_kpi1[f"{col_kpi1}89"].number_format = "#,##0"

    print(f"✅ FILE 10: Khiếu nại T{month} = {knkh_cases}")

    # ==================================================
    # FILE 11 – SỐ TIỀN HOÀN THÀNH (BÁO CÁO KPI MMK)
    # ==================================================
    path_file11 = FILE_CONFIG["file11"]["path"]
    sheet_file11 = "So_tien_HT"

    df_f11 = pd.read_excel(
        path_file11,
        sheet_name=sheet_file11,
        header=0
    )

    # -----------------------------
    # HÀNG 3 = index 2
    # Cột T1 = C = index 2 → + (month - 1)
    # -----------------------------
    row_money_vnd = 2
    col_month_idx = 2 + (month - 1)

    money_vnd = pd.to_numeric(
        df_f11.iloc[row_money_vnd, col_month_idx],
        errors="coerce"
    )

    money_usd = money_vnd / EXCHANGE_RATE if pd.notna(money_vnd) else 0

    print(
        "✅ FILE 11 – SỐ TIỀN HOÀN THÀNH\n"
        f"   VND = {money_vnd:,.0f}\n"
        f"   USD = {money_usd:,.0f} (rate {EXCHANGE_RATE})"
    )
    # ==================================================
    # GHI VÀO 指標４（生産性) – HÀNG 38
    # ==================================================
    ws_kpi4 = wb["指標４（生産性) "]
    col_kpi4 = excel_col(9 + month - 1)  # I = tháng 1

    ws_kpi4[f"{col_kpi4}38"] = money_usd
    ws_kpi4[f"{col_kpi4}38"].number_format = "#,##0"

    print(
        f"✅ 指標４（生産性) | Row 38 | Cột {col_kpi4}\n"
        f"   Giá trị ghi = {money_usd:,.0f} USD"
    )


    wb.save(current_file)
    print("✅ KPI BATCH HOÀN TẤT")

if __name__ == "__main__":
    try:
        run_kpi()
    except Exception as e:
        print(f"❌ KPI FAILED: {e}")